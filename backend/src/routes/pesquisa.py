from flask import Blueprint, request, jsonify, send_file, redirect, Response
from models.registro import Registro, db
from models.obra import Obra
from models.tipo_registro import TipoRegistro
from routes.auth import token_required, obra_access_required
from services.blob_service import blob_service
from sqlalchemy import or_, and_, func
from datetime import datetime
import os
import requests
import mimetypes
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

pesquisa_bp = Blueprint('pesquisa', __name__)


@pesquisa_bp.route('/', methods=['GET'])
@token_required
@obra_access_required
def pesquisa_avancada(current_user):
    try:
        # Parâmetros de busca
        palavra_chave = request.args.get('palavra_chave', '').strip()
        obra_id = request.args.get('obra_id', type=int)
        tipo_registro = request.args.get('tipo_registro', '').strip()
        tipo_registro_id = request.args.get('tipo_registro_id', type=int)
        codigo_numero = request.args.get('codigo_numero', '').strip()
        autor_id = request.args.get('autor_id', type=int)
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        data_registro_inicio = request.args.get('data_registro_inicio')
        data_registro_fim = request.args.get('data_registro_fim')
        # ✅ NOVO: Filtro por classificação
        classificacao_grupo = request.args.get(
            'classificacao_grupo', '').strip()

        # Parâmetros de paginação
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)

        # Query base
        query = Registro.query

        # Aplicar filtros de acesso baseado no usuário
        if current_user.role == 'usuario_padrao':
            query = query.filter_by(obra_id=current_user.obra_id)
        elif obra_id:
            query = query.filter_by(obra_id=obra_id)

        # Filtro por palavra-chave (busca em título e descrição)
        if palavra_chave:
            query = query.filter(
                or_(
                    Registro.titulo.ilike(f'%{palavra_chave}%'),
                    Registro.descricao.ilike(f'%{palavra_chave}%')
                )
            )

        # Filtro por tipo de registro
        if tipo_registro:
            query = query.filter_by(tipo_registro=tipo_registro)

        if tipo_registro_id:
            query = query.filter_by(tipo_registro_id=tipo_registro_id)

        # ✅ NOVO: Filtro por classificação
        if classificacao_grupo:
            query = query.filter_by(classificacao_grupo=classificacao_grupo)

        # Filtro por código/número
        if codigo_numero:
            query = query.filter(
                Registro.codigo_numero.ilike(f'%{codigo_numero}%'))

        # Filtro por autor
        if autor_id:
            query = query.filter_by(autor_id=autor_id)

        # Filtros por data de criação
        if data_inicio:
            try:
                data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                query = query.filter(Registro.created_at >= data_inicio_dt)
            except ValueError:
                return jsonify({'message': 'Formato de data_inicio inválido (use YYYY-MM-DD)'}), 400

        if data_fim:
            try:
                data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                # Adicionar 1 dia para incluir todo o dia final
                data_fim_dt = data_fim_dt.replace(
                    hour=23, minute=59, second=59)
                query = query.filter(Registro.created_at <= data_fim_dt)
            except ValueError:
                return jsonify({'message': 'Formato de data_fim inválido (use YYYY-MM-DD)'}), 400

        # Filtros por data do registro
        if data_registro_inicio:
            try:
                data_registro_inicio_dt = datetime.strptime(
                    data_registro_inicio, '%Y-%m-%d')
                data_registro_fim_dt = data_registro_inicio_dt.replace(
                    hour=23, minute=59, second=59)
                query = query.filter(
                    Registro.data_registro >= data_registro_inicio_dt,
                    Registro.data_registro <= data_registro_fim_dt)
            except ValueError:
                return jsonify({'message': 'Formato de data_registro_inicio inválido (use YYYY-MM-DD)'}), 400

        if data_registro_fim:
            try:
                data_registro_fim_dt = datetime.strptime(
                    data_registro_fim, '%Y-%m-%d')
                data_registro_fim_dt = data_registro_fim_dt.replace(
                    hour=23, minute=59, second=59)
                query = query.filter(
                    Registro.data_registro <= data_registro_fim_dt)
            except ValueError:
                return jsonify({'message': 'Formato de data_registro_fim inválido (use YYYY-MM-DD)'}), 400

        # Ordenação
        ordenacao = request.args.get('ordenacao', 'data_desc')
        if ordenacao == 'data_asc':
            query = query.order_by(Registro.created_at.asc())
        elif ordenacao == 'data_desc':
            query = query.order_by(Registro.created_at.desc())
        elif ordenacao == 'titulo_asc':
            query = query.order_by(Registro.titulo.asc())
        elif ordenacao == 'titulo_desc':
            query = query.order_by(Registro.titulo.desc())
        elif ordenacao == 'data_registro_asc':
            query = query.order_by(Registro.data_registro.asc())
        elif ordenacao == 'data_registro_desc':
            query = query.order_by(Registro.data_registro.desc())
        else:
            query = query.order_by(Registro.created_at.desc())

        # Paginação
        registros_paginados = query.paginate(
            page=page, per_page=per_page, error_out=False
        )

        return jsonify({
            'registros': [registro.to_dict() for registro in registros_paginados.items],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': registros_paginados.total,
                'pages': registros_paginados.pages,
                'has_next': registros_paginados.has_next,
                'has_prev': registros_paginados.has_prev
            },
            'filtros_aplicados': {
                'palavra_chave': palavra_chave,
                'obra_id': obra_id,
                'tipo_registro': tipo_registro,
                'tipo_registro_id': tipo_registro_id,
                'codigo_numero': codigo_numero,
                'autor_id': autor_id,
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'data_registro_inicio': data_registro_inicio,
                'data_registro_fim': data_registro_fim,
                'classificacao_grupo': classificacao_grupo,
                'ordenacao': ordenacao
            }
        }), 200

    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500


@pesquisa_bp.route('/filtros', methods=['GET'])
@token_required
@obra_access_required
def get_filtros_disponiveis(current_user):
    try:
        # Obras disponíveis
        obras = []
        if current_user.role == 'administrador':
            obras = Obra.query.all()
        else:
            obra = Obra.query.get(current_user.obra_id)
            if obra:
                obras = [obra]

        # Tipos de registro disponíveis
        tipos_registro = TipoRegistro.query.filter_by(ativo=True).all()

        # Tipos de registro únicos nos registros existentes
        query = db.session.query(Registro.tipo_registro).distinct()
        if current_user.role == 'usuario_padrao':
            query = query.filter_by(obra_id=current_user.obra_id)

        tipos_registro_existentes = [tipo[0]
                                     for tipo in query.all() if tipo[0]]

        # ✅ NOVO: Classificações disponíveis
        classificacoes_query = db.session.query(
            Registro.classificacao_grupo).distinct()
        if current_user.role == 'usuario_padrao':
            classificacoes_query = classificacoes_query.filter_by(
                obra_id=current_user.obra_id)

        classificacoes_existentes = [cls[0]
                                     for cls in classificacoes_query.all() if cls[0]]

        # Autores (usuários que criaram registros)
        from models.user import User
        autores_query = db.session.query(User.id, User.username, User.email)\
            .join(Registro, User.id == Registro.autor_id).distinct()

        if current_user.role == 'usuario_padrao':
            autores_query = autores_query.filter(
                Registro.obra_id == current_user.obra_id)

        autores = autores_query.all()

        # Faixas de data disponíveis
        data_query = db.session.query(
            func.min(Registro.created_at).label('data_min'),
            func.max(Registro.created_at).label('data_max'),
            func.min(Registro.data_registro).label('data_registro_min'),
            func.max(Registro.data_registro).label('data_registro_max')
        )

        if current_user.role == 'usuario_padrao':
            data_query = data_query.filter_by(obra_id=current_user.obra_id)

        datas = data_query.first()

        return jsonify({
            'obras': [obra.to_dict() for obra in obras],
            'tipos_registro': [tipo.to_dict() for tipo in tipos_registro],
            'tipos_registro_existentes': tipos_registro_existentes,
            # ✅ NOVO: Classificações nos filtros
            'classificacoes_existentes': classificacoes_existentes,
            'autores': [
                {'id': autor_id, 'username': username, 'email': email}
                for autor_id, username, email in autores
            ],
            'faixas_data': {
                'criacao_min': datas.data_min.isoformat() if datas.data_min else None,
                'criacao_max': datas.data_max.isoformat() if datas.data_max else None,
                'registro_min': datas.data_registro_min.isoformat() if datas.data_registro_min else None,
                'registro_max': datas.data_registro_max.isoformat() if datas.data_registro_max else None
            },
            'opcoes_ordenacao': [
                {'value': 'data_desc',
                    'label': 'Data de Criação (Mais Recente)'},
                {'value': 'data_asc',
                    'label': 'Data de Criação (Mais Antiga)'},
                {'value': 'data_registro_desc',
                    'label': 'Data do Registro (Mais Recente)'},
                {'value': 'data_registro_asc',
                    'label': 'Data do Registro (Mais Antiga)'},
                {'value': 'titulo_asc', 'label': 'Título (A-Z)'},
                {'value': 'titulo_desc', 'label': 'Título (Z-A)'}
            ]
        }), 200

    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500


# ✅ NOVO: Exportação Excel
@pesquisa_bp.route('/exportar', methods=['POST'])
@token_required
@obra_access_required
def exportar_resultados(current_user):
    try:
        # Receber os mesmos filtros da pesquisa
        data = request.get_json()

        print(
            f"📊 EXPORTAÇÃO: Iniciando exportação para usuário {current_user.username}")
        print(f"   - Filtros recebidos: {data}")

        # Aplicar os mesmos filtros da pesquisa avançada (sem paginação)
        palavra_chave = data.get('palavra_chave', '').strip()
        obra_id = data.get('obra_id')
        tipo_registro_id = data.get('tipo_registro_id')
        classificacao_grupo = data.get('classificacao_grupo', '').strip()
        data_registro_inicio = data.get('data_registro_inicio')

        # Query base
        query = Registro.query

        # Aplicar filtros de acesso baseado no usuário
        if current_user.role == 'usuario_padrao':
            query = query.filter_by(obra_id=current_user.obra_id)
        elif obra_id and obra_id != '0':
            query = query.filter_by(obra_id=int(obra_id))

        # Aplicar filtros
        if palavra_chave:
            query = query.filter(
                or_(
                    Registro.titulo.ilike(f'%{palavra_chave}%'),
                    Registro.descricao.ilike(f'%{palavra_chave}%')
                )
            )

        if tipo_registro_id and tipo_registro_id != '0':
            query = query.filter_by(tipo_registro_id=int(tipo_registro_id))

        if classificacao_grupo:
            query = query.filter_by(classificacao_grupo=classificacao_grupo)

        if data_registro_inicio:
            try:
                data_dt = datetime.strptime(data_registro_inicio, '%Y-%m-%d')
                query = query.filter(Registro.data_registro >= data_dt)
            except ValueError:
                pass

        # Buscar todos os registros (sem paginação)
        registros = query.order_by(Registro.created_at.desc()).all()

        print(f"📊 EXPORTAÇÃO: {len(registros)} registros encontrados")

        if not registros:
            return jsonify({'message': 'Nenhum registro encontrado para exportar'}), 404

        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros GEDO CIMCOP"

        # Definir cabeçalhos
        headers = [
            'Data do Registro',
            'Obra',
            'Tipo de Registro',
            'Classificação',
            'Título',
            'Autor',
            'Descrição'
        ]

        # Estilizar cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Adicionar dados
        for row, registro in enumerate(registros, 2):
            ws.cell(row=row, column=1, value=registro.data_registro.strftime(
                '%d/%m/%Y') if registro.data_registro else '')
            ws.cell(row=row, column=2,
                    value=registro.obra.nome if registro.obra else '')
            ws.cell(row=row, column=3, value=registro.tipo_registro or '')
            ws.cell(row=row, column=4,
                    value=f"{registro.classificacao_grupo} > {registro.classificacao_subgrupo}" if registro.classificacao_grupo and registro.classificacao_subgrupo else '')
            ws.cell(row=row, column=5, value=registro.titulo or '')
            ws.cell(row=row, column=6,
                    value=registro.autor.username if registro.autor else '')
            ws.cell(row=row, column=7, value=registro.descricao or '')

        # Ajustar largura das colunas
        column_widths = [15, 25, 20, 30, 30, 15, 50]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(
                row=1, column=col).column_letter].width = width

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'registros_gedo_cimcop_{timestamp}.xlsx'

        print(f"✅ EXPORTAÇÃO: Arquivo Excel criado: {filename}")

        return Response(
            output.getvalue(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Cache-Control': 'no-cache, no-store, must-revalidate',
                'Pragma': 'no-cache',
                'Expires': '0'
            }
        )

    except Exception as e:
        print(f"❌ EXPORTAÇÃO: Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'message': f'Erro interno na exportação: {str(e)}'}), 500


@pesquisa_bp.route('/<int:registro_id>/download', methods=['GET'])
@token_required
@obra_access_required
def download_anexo(current_user, registro_id):
    try:
        print(f"🔽 DOWNLOAD: Iniciando download do registro {registro_id}")
        print(f"   - Usuário: {current_user.username} (ID: {current_user.id})")

        registro = Registro.query.get(registro_id)
        if not registro:
            print(f"❌ DOWNLOAD: Registro {registro_id} não encontrado")
            return jsonify({'message': 'Registro não encontrado'}), 404

        print(f"✅ DOWNLOAD: Registro {registro_id} encontrado")
        print(f"   - Título: {registro.titulo}")
        print(f"   - Nome original: {registro.nome_arquivo_original}")
        print(f"   - Formato: {registro.formato_arquivo}")

        # Verificar permissões
        if current_user.role == 'usuario_padrao' and registro.obra_id != current_user.obra_id:
            print(
                f"❌ DOWNLOAD: Acesso negado - usuário obra {current_user.obra_id} != registro obra {registro.obra_id}")
            return jsonify({'message': 'negado a este registro'}), 403

        print(f"✅ DOWNLOAD: Permissões OK")
        print(f"   - Tem blob_url: {bool(registro.blob_url)}")
        print(f"   - Tem caminho_anexo: {bool(registro.caminho_anexo)}")

        # ← CORRIGIDO: Melhor detecção de tipo e nome do arquivo
        if registro.blob_url:
            try:
                print(f"🔗 DOWNLOAD: Fazendo proxy do Vercel Blob")
                print(f"   - URL: {registro.blob_url}")

                # Headers para requisição
                headers = {
                    'User-Agent': 'GEDO-CIMCOP/1.0',
                    'Accept': '*/*'
                }

                print("📡 DOWNLOAD: Fazendo requisição para Vercel Blob...")
                response = requests.get(
                    registro.blob_url, headers=headers, stream=True, timeout=60)
                response.raise_for_status()

                print(
                    f"✅ DOWNLOAD: Resposta do Vercel Blob: {response.status_code}")
                print(
                    f"   - Content-Type original: {response.headers.get('content-type')}")
                print(
                    f"   - Content-Length: {response.headers.get('content-length')}")

                # ← CORREÇÃO CRÍTICA: Melhor detecção de Content-Type e nome do arquivo

                # 1. Determinar extensão do arquivo
                file_extension = None
                if registro.formato_arquivo:
                    file_extension = registro.formato_arquivo.lower()
                elif registro.nome_arquivo_original and '.' in registro.nome_arquivo_original:
                    file_extension = registro.nome_arquivo_original.rsplit('.', 1)[
                        1].lower()

                print(f"📎 DOWNLOAD: Extensão detectada: {file_extension}")

                # 2. Determinar Content-Type correto baseado na extensão
                content_type_map = {
                    'pdf': 'application/pdf',
                    'doc': 'application/msword',
                    'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                    'xls': 'application/vnd.ms-excel',
                    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'ppt': 'application/vnd.ms-powerpoint',
                    'pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
                    'png': 'image/png',
                    'jpg': 'image/jpeg',
                    'jpeg': 'image/jpeg',
                    'gif': 'image/gif',
                    'bmp': 'image/bmp',
                    'tiff': 'image/tiff',
                    'txt': 'text/plain',
                    'csv': 'text/csv',
                    'zip': 'application/zip',
                    'rar': 'application/x-rar-compressed',
                    '7z': 'application/x-7z-compressed',
                    'mp4': 'video/mp4',
                    'avi': 'video/x-msvideo',
                    'mp3': 'audio/mpeg',
                    'wav': 'audio/wav'
                }

                # Usar Content-Type baseado na extensão
                if file_extension and file_extension in content_type_map:
                    content_type = content_type_map[file_extension]
                else:
                    # Tentar usar o Content-Type da resposta
                    content_type = response.headers.get(
                        'content-type', 'application/octet-stream')
                    # Se for genérico, usar mimetypes
                    if content_type == 'application/octet-stream' and file_extension:
                        guessed_type = mimetypes.guess_type(
                            f"file.{file_extension}")[0]
                        if guessed_type:
                            content_type = guessed_type

                print(f"📎 DOWNLOAD: Content-Type final: {content_type}")

                # 3. Determinar nome do arquivo com extensão correta
                if registro.nome_arquivo_original:
                    filename = registro.nome_arquivo_original
                    # Garantir que tem a extensão correta
                    if file_extension and not filename.lower().endswith(f'.{file_extension}'):
                        # Se o nome não tem extensão, adicionar
                        if '.' not in filename:
                            filename = f"{filename}.{file_extension}"
                        # Se tem extensão diferente, substituir
                        else:
                            base_name = filename.rsplit('.', 1)[0]
                            filename = f"{base_name}.{file_extension}"
                else:
                    # Nome padrão com extensão
                    if file_extension:
                        filename = f"anexo_{registro_id}.{file_extension}"
                    else:
                        filename = f"anexo_{registro_id}"

                print(f"📎 DOWNLOAD: Nome do arquivo final: {filename}")

                # 4. Criar resposta streaming com headers corretos
                def generate():
                    try:
                        for chunk in response.iter_content(chunk_size=8192):
                            if chunk:
                                yield chunk
                    except Exception as e:
                        print(f"❌ DOWNLOAD: Erro no streaming: {str(e)}")
                        raise

                print("🚀 DOWNLOAD: Iniciando streaming do arquivo...")

                # ← CORREÇÃO: Headers mais específicos para forçar download correto
                response_headers = {
                    'Content-Type': content_type,
                    'Content-Disposition': f'attachment; filename="{filename}"',
                    'Cache-Control': 'no-cache, no-store, must-revalidate',
                    'Pragma': 'no-cache',
                    'Expires': '0',
                    'X-Content-Type-Options': 'nosniff'
                }

                # Adicionar Content-Length se disponível
                content_length = response.headers.get('content-length')
                if content_length:
                    response_headers['Content-Length'] = content_length

                print(f"📋 DOWNLOAD: Headers da resposta: {response_headers}")

                return Response(
                    generate(),
                    headers=response_headers
                )

            except requests.RequestException as e:
                print(f"❌ DOWNLOAD: Erro ao baixar do Vercel Blob: {str(e)}")
                return jsonify({'message': f'Erro ao acessar arquivo no storage: {str(e)}'}), 500
            except Exception as e:
                print(f"❌ DOWNLOAD: Erro geral no Blob: {str(e)}")
                return jsonify({'message': f'Erro interno no download: {str(e)}'}), 500

        # Fallback para sistema antigo
        if not registro.caminho_anexo:
            print("❌ DOWNLOAD: Registro não possui anexo")
            return jsonify({'message': 'Este registro não possui anexo'}), 404

        print(f"📁 DOWNLOAD: Usando sistema local: {registro.caminho_anexo}")
        if not os.path.exists(registro.caminho_anexo):
            print(
                f"❌ DOWNLOAD: Arquivo local não encontrado: {registro.caminho_anexo}")
            return jsonify({'message': 'Arquivo não encontrado no servidor'}), 404

        print("✅ DOWNLOAD: Enviando arquivo local")

        # ← CORREÇÃO: Melhor detecção para arquivos locais também
        filename = registro.nome_arquivo_original or f'anexo_{registro_id}'
        if registro.formato_arquivo and not filename.lower().endswith(f'.{registro.formato_arquivo}'):
            if '.' not in filename:
                filename = f"{filename}.{registro.formato_arquivo}"

        # Detectar mimetype para arquivo local
        mimetype = 'application/octet-stream'
        if registro.formato_arquivo:
            guessed_type = mimetypes.guess_type(
                f"file.{registro.formato_arquivo}")[0]
            if guessed_type:
                mimetype = guessed_type

        return send_file(
            registro.caminho_anexo,
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype
        )

    except Exception as e:
        print(f"❌ DOWNLOAD: Erro geral: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500


@pesquisa_bp.route('/<int:registro_id>/debug', methods=['GET'])
@token_required
def debug_registro(current_user, registro_id):
    """Debug de registro para troubleshooting"""
    try:
        print(f"🔍 DEBUG: Buscando registro ID {registro_id}")

        registro = Registro.query.get(registro_id)
        if not registro:
            print(f"❌ DEBUG: Registro {registro_id} não encontrado")
            return jsonify({
                'message': f'Registro {registro_id} não encontrado',
                'exists': False
            }), 404

        print(f"✅ DEBUG: Registro {registro_id} encontrado")
        print(f"   - Título: {registro.titulo}")
        print(f"   - Tem blob_url: {bool(registro.blob_url)}")
        print(f"   - Tem caminho_anexo: {bool(registro.caminho_anexo)}")
        print(f"   - Blob URL: {registro.blob_url}")
        print(f"   - Caminho anexo: {registro.caminho_anexo}")
        print(f"   - Formato arquivo: {registro.formato_arquivo}")
        print(f"   - Nome original: {registro.nome_arquivo_original}")

        return jsonify({
            'message': 'Registro encontrado',
            'exists': True,
            'registro': {
                'id': registro.id,
                'titulo': registro.titulo,
                'tem_blob_url': bool(registro.blob_url),
                'tem_caminho_anexo': bool(registro.caminho_anexo),
                'blob_url': registro.blob_url,
                'caminho_anexo': registro.caminho_anexo,
                'nome_arquivo_original': registro.nome_arquivo_original,
                'formato_arquivo': registro.formato_arquivo,
                'tem_anexo': bool(registro.blob_url or registro.caminho_anexo)
            }
        }), 200

    except Exception as e:
        print(f"❌ DEBUG: Erro ao buscar registro {registro_id}: {str(e)}")
        return jsonify({
            'message': f'Erro interno: {str(e)}',
            'exists': False
        }), 500


# ✅ NOVO: Visualização completa de registro
@pesquisa_bp.route('/<int:registro_id>/visualizar', methods=['GET'])
@token_required
@obra_access_required
def visualizar_registro(current_user, registro_id):
    """Visualização completa de um registro"""
    try:
        print(f"👁️ VISUALIZAÇÃO: Buscando registro ID {registro_id}")

        registro = Registro.query.get(registro_id)
        if not registro:
            return jsonify({'message': 'Registro não encontrado'}), 404

        # Verificar permissões
        if current_user.role == 'usuario_padrao' and registro.obra_id != current_user.obra_id:
            return jsonify({'message': 'Acesso negado a este registro'}), 403

        print(f"✅ VISUALIZAÇÃO: Registro encontrado: {registro.titulo}")

        return jsonify({
            'registro': registro.to_dict()
        }), 200

    except Exception as e:
        print(f"❌ VISUALIZAÇÃO: Erro: {str(e)}")
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500
